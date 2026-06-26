import json
import time
import random
import os
import sys
import urllib.request
import websocket

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

CDP_PORT = 9222
SAVE_INTERVAL = 100  # 每100条自动保存一次
DEAD_TIMEOUT = 60    # 连续60秒无新数据才判定到底


def get_pages():
    """获取所有页面列表"""
    url = f"http://localhost:{CDP_PORT}/json"
    resp = urllib.request.urlopen(url, timeout=5)
    pages = json.loads(resp.read().decode("utf-8"))
    return [p for p in pages if p.get("type") == "page"]


def select_jd_page(pages):
    """选择京东商品页"""
    jd_pages = [p for p in pages if "jd.com" in p.get("url", "")]
    if not jd_pages:
        return pages[0]["id"] if pages else None
    if len(jd_pages) == 1:
        return jd_pages[0]["id"]
    print("  检测到多个京东页面:")
    for i, p in enumerate(jd_pages):
        print(f"    {i+1}. {p['url'][:80]}")
    return jd_pages[0]["id"]


def cdp_eval(ws, js):
    """执行JS并返回结果"""
    msg_id = random.randint(1, 99999)
    try:
        ws.send(json.dumps({
            "id": msg_id,
            "method": "Runtime.evaluate",
            "params": {"expression": js, "awaitPromise": True, "returnByValue": True}
        }))
        for _ in range(50):
            raw = ws.recv()
            r = json.loads(raw)
            if r.get("id") == msg_id:
                result = r.get("result", {}).get("result", {})
                if result.get("type") == "string":
                    return result.get("value")
                return result.get("value")
    except Exception:
        pass
    return None


def check_risk_control(ws):
    """检测是否触发风控验证 - 只检测真正的验证弹窗"""
    js = """
    (function(){
        var selectors = [
            '.JDJRV-bigimg',
            '.jdc-dialog',
            '[class*="verify"]',
            '[class*="captcha"]',
            'iframe[src*="verify"]',
            'iframe[src*="captcha"]'
        ];
        for(var i=0; i<selectors.length; i++){
            var el = document.querySelector(selectors[i]);
            if(el && el.offsetParent !== null) return 'blocked';
        }
        return 'ok';
    })()
    """
    return cdp_eval(ws, js)


def get_page_title(ws):
    """获取当前页面标题"""
    return cdp_eval(ws, "document.title") or ""


def extract_comments(ws):
    """从当前DOM提取可见评论"""
    js = """
    (function(){
        var items = document.querySelectorAll('[class*="_listItem"]');
        var arr = [];
        items.forEach(function(item){
            var c = {};
            var nick = item.querySelector('.jdc-pc-rate-card-nick');
            c.member = nick ? nick.textContent.trim() : '';

            var stars = item.querySelectorAll('.jdc-pc-rate-card-star .jdc-icon-star-full, [class*="star"][class*="full"]');
            c.star = stars.length > 0 ? stars.length + '星' : '5星';

            var contentEl = item.querySelector('.jdc-pc-rate-card-content');
            if(contentEl){
                var clone = contentEl.cloneNode(true);
                clone.querySelectorAll('[class*="info"]').forEach(function(el){el.remove()});
                c.content = clone.textContent.trim();
            } else { c.content = ''; }

            var infoEl = item.querySelector('.jdc-pc-rate-card-info-left');
            if(infoEl){
                var t = infoEl.textContent.trim();
                var m = t.match(/^(\d{4}-\d{2}-\d{2}|\d{2}-\d{2})/);
                c.date = m ? m[1] : '';
                c.sku = t.replace(/^\d{4}-\d{2}-\d{2}|^\d{2}-\d{2}/, '').trim();
            } else { c.date=''; c.sku=''; }

            var likeEl = item.querySelector('[class*="like"] span, [class*="zan"] span');
            c.likes = likeEl ? likeEl.textContent.trim() : '';

            var replyCount = item.querySelectorAll('[class*="reply"], [class*="comment"]').length;
            c.replies = replyCount > 0 ? String(replyCount) : '';

            var imgs = item.querySelectorAll('img[src*="jfs"]');
            c.imageCount = imgs.length;

            if(c.member || c.content) arr.push(c);
        });
        return JSON.stringify(arr);
    })()
    """
    val = cdp_eval(ws, js)
    if val:
        try:
            return json.loads(val)
        except Exception:
            return []
    return []


def save_to_xlsx(all_comments, out_path, page_title="", good_rate=""):
    try:
        from openpyxl import Workbook
    except ImportError:
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        from openpyxl import Workbook
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    headers = ["会员", "级别", "评价星级", "评价内容", "时间",
               "点赞数", "评论数", "追评时间", "追评内容",
               "商品属性", "页面标题", "好评度", "评价关键词"]
    for col, h in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=h)
    for idx, c in enumerate(all_comments, 1):
        row = idx + 1
        sheet.cell(row=row, column=1, value=c.get("member", ""))
        sheet.cell(row=row, column=2, value=c.get("level", ""))
        sheet.cell(row=row, column=3, value=c.get("star", "5星"))
        sheet.cell(row=row, column=4, value=c.get("content", ""))
        sheet.cell(row=row, column=5, value=c.get("date", ""))
        sheet.cell(row=row, column=6, value=c.get("likes", ""))
        sheet.cell(row=row, column=7, value=c.get("replies", ""))
        sheet.cell(row=row, column=8, value="")
        sheet.cell(row=row, column=9, value="")
        sheet.cell(row=row, column=10, value=c.get("sku", ""))
        sheet.cell(row=row, column=11, value=page_title)
        sheet.cell(row=row, column=12, value=good_rate)
        sheet.cell(row=row, column=13, value="")
    wb.save(out_path)
    print(f"\n  [保存] {len(all_comments)} 条评论 -> {out_path}")


def scroll_container(ws, distance=800):
    """滚动评论容器 - 尝试多个可能的容器"""
    cdp_eval(ws, f"""
    (function(){{
        var containers = [
            document.querySelector('[class*="_rateListContainer"]'),
            document.querySelector('[class*="_rateListBox"]'),
            document.querySelector('.jdc-page-overlay'),
            document.querySelector('[class*="rate"] [class*="list"]')
        ];
        for(var i=0; i<containers.length; i++){{
            var c = containers[i];
            if(c && c.scrollHeight > c.clientHeight){{
                c.scrollTop += {distance};
                return;
            }}
        }}
    }})()
    """)


def main():
    print("=" * 50)
    print("京东评论爬虫 - 持续抓取 / 风控保存 / 切换商品追加")
    print(f"CDP端口: {CDP_PORT}")
    print(f"到底判定: 连续{DEAD_TIMEOUT}秒无新评论")
    print("=" * 50)

    out_dir = os.path.dirname(os.path.abspath(__file__))
    all_comments = []
    seen = set()

    while True:
        # 连接Chrome
        try:
            pages = get_pages()
        except Exception as e:
            print(f"[错误] 无法连接Chrome CDP端口{CDP_PORT}: {e}")
            return

        target_id = select_jd_page(pages)
        if not target_id:
            print("[错误] 没有找到可用页面")
            return

        ws = websocket.create_connection(
            f"ws://localhost:{CDP_PORT}/devtools/page/{target_id}",
            timeout=30, suppress_origin=True
        )

        page_title = get_page_title(ws)
        print(f"\n当前页面: {page_title[:60]}")

        # 获取好评度
        good_rate = cdp_eval(ws, """
        (function(){
            var el = document.querySelector('[class*="percent"], [class*="good-rate"]');
            return el ? el.textContent.trim() : '';
        })()
        """) or ""

        # 打开评论弹窗
        print("  打开评论弹窗...")
        cdp_eval(ws, "document.querySelector('.comment-root').scrollIntoView()")
        time.sleep(1)
        cdp_eval(ws, "document.querySelector('.all-btn').click()")
        time.sleep(3)

        # 验证弹窗
        check = cdp_eval(ws, """
        (function(){
            var o = document.querySelector('[class*="_rateListBox"],.jdc-page-overlay');
            return (o && getComputedStyle(o).display !== 'none') ? 'open' : 'closed';
        })()
        """)
        if check != 'open':
            print("  [警告] 弹窗未打开，尝试直接抓取...")

        # 开始滚动抓取
        print("  开始抓取评论（只有风控或60秒无新数据才会停）...")
        batch_start = len(all_comments)
        last_new_time = time.time()  # 上次获得新评论的时间
        loop_count = 0

        while True:
            loop_count += 1

            # 风控检测（每5轮检测一次，减少开销）
            if loop_count % 5 == 0:
                risk = check_risk_control(ws)
                if risk == 'blocked':
                    print(f"\n  [风控] 触发验证！已抓取 {len(all_comments)} 条")
                    out_path = os.path.join(out_dir, f"comments_{len(all_comments)}.xlsx")
                    save_to_xlsx(all_comments, out_path, page_title, good_rate)
                    ws.close()
                    print("\n" + "=" * 50)
                    print("请完成验证，然后切换到新商品页面")
                    input("切换好后按回车继续...")
                    print("=" * 50)
                    break

            # 提取评论
            comments = extract_comments(ws)
            new_count = 0
            for c in comments:
                key = c.get('member', '') + '|' + c.get('content', '')[:50]
                if key not in seen and key != '|':
                    seen.add(key)
                    all_comments.append(c)
                    new_count += 1

            if new_count > 0:
                last_new_time = time.time()

            current_total = len(all_comments)
            batch_count = current_total - batch_start
            elapsed_no_new = int(time.time() - last_new_time)
            print(f"\r  本轮: {batch_count} | 总计: {current_total} | 静默: {elapsed_no_new}s", end="", flush=True)

            # 定期保存
            if current_total > 0 and current_total % SAVE_INTERVAL < 3 and new_count > 0:
                out_path = os.path.join(out_dir, f"comments_{current_total}.xlsx")
                save_to_xlsx(all_comments, out_path, page_title, good_rate)

            # 唯一的到底判定：连续DEAD_TIMEOUT秒没有任何新评论
            if time.time() - last_new_time > DEAD_TIMEOUT:
                print(f"\n  [完成] 连续{DEAD_TIMEOUT}秒无新评论，本商品抓取结束")
                out_path = os.path.join(out_dir, f"comments_{len(all_comments)}.xlsx")
                save_to_xlsx(all_comments, out_path, page_title, good_rate)
                ws.close()
                print("\n需要继续抓取下一个商品吗？")
                choice = input("切换商品后按回车继续，输入 q 退出: ").strip()
                if choice.lower() == 'q':
                    print(f"\n最终结果: {len(all_comments)} 条评论")
                    print(f"文件: {out_path}")
                    return
                break

            # 无论是否有新数据，都继续滚动
            scroll_distance = random.randint(500, 1000)
            scroll_container(ws, scroll_distance)
            time.sleep(random.uniform(1.5, 3.5))


if __name__ == "__main__":
    main()
